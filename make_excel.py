import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. GENERATE LARGE SAMPLE DATA
# Sheet 1: Inventory (Warehouse Catalog)
inventory_data = {
    "Item ID": [f"INV-{i:03d}" for i in range(1, 16)],
    "Item Name": [
        "Fruit Up Toffee", "Jelly Center Filled Gum", "Choco Bliss Bar",
        "Mango Tang Candy", "Sour Punk Strips", "Bubble Bill Gum",
        "Cream Biscuit Pack", "Salted Potato Chips", "Zeera Plus Biscuit",
        "Wafers Chocolate", "Peanut Chilli Nimko", "Mix Dry Fruits Box",
        "Lollipop Strawberry", "Mint Fresh Strips", "Milky Toffee Classic"
    ],
    "Category": [
        "Confectionery", "Confectionery", "Chocolates",
        "Confectionery", "Confectionery", "Confectionery",
        "Snacks", "Snacks", "Snacks",
        "Snacks", "Snacks", "Snacks",
        "Confectionery", "Confectionery", "Confectionery"
    ],
    "Packing Size": [
        "12x50 Box", "24x30 Pack", "50pcs Carton",
        "12x50 Box", "20pcs Pack", "40pcs Box",
        "24 Packets", "50 Packets", "24 Packets",
        "12 Packets", "1kg Bag", "500g Box",
        "100pcs Jar", "50pcs Pack", "12x50 Box"
    ],
    "Unit": ["Jars", "Packets", "Cartons", "Jars", "Packets", "Jars", "Cartons", "Cartons", "Cartons", "Cartons", "Kg",
             "Pcs", "Jars", "Packets", "Jars"],
    "Current Stock": [350, 80, 500, 45, 120, 600, 30, 400, 250, 15, 90, 120, 40, 750, 180],
    "Safety Stock": [100, 100, 150, 50, 50, 200, 50, 100, 80, 30, 40, 30, 50, 150, 100]
}
df_inventory = pd.DataFrame(inventory_data)

# Sheet 2: Order Booking Log (Large Sample Data)
orders_data = {
    "Order ID": [f"ORD-{i:04d}" for i in range(1, 13)],
    "Client Name": [
        "Waqas Shakeel", "Zain Ali Traders", "Bilal Food Agency", "Waqas Shakeel",
        "Alpha Super Mart", "Zain Ali Traders", "Bilal Food Agency", "Waqas Shakeel",
        "Kashif Confectionery", "Alpha Super Mart", "Zain Ali Traders", "Mian Distributors"
    ],
    "Product Name": [
        "Fruit Up Toffee", "Jelly Center Filled Gum", "Choco Bliss Bar", "Mango Tang Candy",
        "Cream Biscuit Pack", "Wafers Chocolate", "Salted Potato Chips", "Jelly Center Filled Gum",
        "Lollipop Strawberry", "Zeera Plus Biscuit", "Fruit Up Toffee", "Milky Toffee Classic"
    ],
    "Quantity Ordered": [400, 100, 150, 80, 60, 45, 200, 350, 120, 90, 50, 300],
    "Fulfillment Status": [
        "Awaiting Production", "Awaiting Production", "Fulfilled", "Awaiting Production",
        "Awaiting Production", "Awaiting Production", "Fulfilled", "Fulfilled",
        "Awaiting Production", "Fulfilled", "Fulfilled", "Awaiting Production"
    ],
    "Date/Time": [
        "2026-07-10 17:07", "2026-07-10 17:06", "2026-07-09 11:20", "2026-07-09 14:45",
        "2026-07-08 09:15", "2026-07-08 10:30", "2026-07-07 16:00", "2026-07-07 16:44",
        "2026-07-06 12:10", "2026-07-05 15:30", "2026-07-05 16:15", "2026-07-04 11:00"
    ]
}
df_orders = pd.DataFrame(orders_data)

# 2. WRITE TO EXCEL FILE WITH ADVANCED FORMATTING
file_name = "Factory_Management_System.xlsx"
with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    df_inventory.to_excel(writer, sheet_name="Inventory & Stock", index=False)
    df_orders.to_excel(writer, sheet_name="Order Booking Log", index=False)

# 3. APPLY EXCEL FORMULAS, STYLES, FILTERS, AND ALERTS
wb = openpyxl.load_workbook(file_name)

# Color Palettes & Styles
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Corporate Blue
low_stock_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Soft Red
low_stock_font = Font(name="Calibri", size=11, bold=True, color="9C0006")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Soft Green
ok_font = Font(name="Calibri", size=11, color="006100")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# ---- FORMATTING SHEET 1: INVENTORY & STOCK ----
ws1 = wb["Inventory & Stock"]
ws1.cell(row=1, column=8, value="Status / Alert").font = header_font  # Add 8th Column Header

# Enable Excel Filter Dropdowns automatically
ws1.auto_filter.ref = f"A1:H{len(df_inventory) + 1}"

for row_idx in range(2, len(df_inventory) + 3):
    # Dynamic formula placement for column H (Status)
    status_cell = ws1.cell(row=row_idx, column=8)
    status_cell.value = f'=IF(F{row_idx}<=G{row_idx}, "🚨 Low Stock / Produce", "✅ OK")'

    # Conditional Styling for Alerts based on value
    current_stock = ws1.cell(row=row_idx, column=6).value
    safety_stock = ws1.cell(row=row_idx, column=7).value
    if current_stock is not None and safety_stock is not None:
        if current_stock <= safety_stock:
            status_cell.fill = low_stock_fill
            status_cell.font = low_stock_font
        else:
            status_cell.fill = ok_fill
            status_cell.font = ok_font

# ---- FORMATTING SHEET 2: ORDER BOOKING LOG ----
ws2 = wb["Order Booking Log"]
ws2.auto_filter.ref = f"A1:F{len(df_orders) + 1}"

# Soft color coding for status column
for row_idx in range(2, len(df_orders) + 2):
    status_cell = ws2.cell(row=row_idx, column=5)
    if status_cell.value == "Awaiting Production":
        status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Soft Yellow
        status_cell.font = Font(name="Calibri", size=11, color="B2A100", bold=True)
    else:
        status_cell.fill = ok_fill
        status_cell.font = ok_font

# ---- GENERAL BEAUTIFICATION FOR ALL SHEETS ----
for ws in [ws1, ws2]:
    # Style Header Rows
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style Data Rows, Alignments, and Grid Borders
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            elif cell.column in [1, 5, 6, 8]:  # IDs, Statuses, Units, Dates
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Auto-fit Column Widths based on data length so nothing looks cropped
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save Workbook
wb.save(file_name)
print(f"🎉 Mubarak ho! Super clean file '{file_name}' create ho chuki hai.")